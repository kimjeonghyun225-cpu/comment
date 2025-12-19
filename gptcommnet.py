# -*- coding: utf-8 -*-
# Streamlit 앱: QA 결과 자동 코멘트 생성기 (최종)

import os, re, io, time, unicodedata
from contextlib import contextmanager
from typing import List, Dict, Any, Optional

import pandas as pd
import openpyxl
import streamlit as st
from dotenv import load_dotenv
from openai import OpenAI

from qa_patch_module import (
    find_test_sheet_candidates,
    extract_comments_as_dataframe_dual,
    enrich_with_column_comments,
    self_check,
    parse_llm_json,
    build_system_prompt,
    build_user_prompt,
    write_excel_report,
    load_threaded_comments_map_from_bytes,
)

# ==============================
# 환경/초기 설정
# ==============================
load_dotenv()

st.set_page_config(
    page_title="호환성 QA 자동 리포트",
    layout="wide",
)

# 세션 초기화 버튼: 이전 업로드/상태를 완전히 비우고 새로 시작
if st.sidebar.button("🔄 세션 초기화 / 새 파일 분석"):
    for k in list(st.session_state.keys()):
        del st.session_state[k]
    st.experimental_rerun()

@contextmanager
def step_status(msg: str):
    st.write(f"### ⏱ {msg}")
    with st.spinner(msg + "..."):
        start = time.time()
        try:
            yield
        finally:
            dt = time.time() - start
            st.write(f"✅ 완료 ({dt:0.1f}s) - {msg}")

def diag_dump(title: str, obj):
    """디버그용 덤프(필요 시만 st.write)."""
    with st.expander(f"🔍 {title}", expanded=False):
        if isinstance(obj, (pd.DataFrame, pd.Series)):
            st.dataframe(obj)
        else:
            st.write(obj)

# ==============================
# OpenAI 클라이언트
# ==============================
api_key = os.getenv("OPENAI_API_KEY", "").strip()
if not api_key:
    st.error("❌ OpenAI API 키가 없습니다. st.secrets 또는 .env에 OPENAI_API_KEY를 설정하세요.")
    st.stop()

client = OpenAI(api_key=api_key)

# ==============================
# UI: 파일 업로드
# ==============================
st.title("호환성 QA 자동 리포트 생성기")

uploaded = st.file_uploader("ex) KP 4.2.0 Build CO QA Report_In.xlsx 파일을 업로드하세요.", type=["xlsx"])
if not uploaded:
    st.stop()

data = uploaded.read()
xls = io.BytesIO(data)

# ==============================
# 1) 테스트 시트 자동 탐지
# ==============================
with step_status("테스트 시트 자동 탐지"):
    try:
        wb = openpyxl.load_workbook(xls, data_only=True)
        sheet_names = wb.sheetnames
        st.write("📄 감지된 시트:", ", ".join(sheet_names))
    except Exception as e:
        st.error(f"엑셀을 열 수 없습니다: {e}")
        st.stop()

    test_candidates = find_test_sheet_candidates(sheet_names)
    st.write("🧪 테스트 시트 후보:", ", ".join(test_candidates) if test_candidates else "(없음)")

    test_sheets_selected = st.multiselect(
        "테스트 결과(Compatibility Test 등)가 포함된 시트를 선택하세요.",
        options=sheet_names,
        default=test_candidates or sheet_names,
    )
    if not test_sheets_selected:
        st.error("테스트 시트를 최소 1개 이상 선택해야 합니다.")
        st.stop()

# ==============================
# 2) 스펙 시트(디바이스 리스트) 선택
# ==============================
with step_status("스펙(디바이스 리스트) 시트 후보 탐지"):
    spec_candidates = []
    for s in sheet_names:
        s_norm = s.lower()
        if any(k in s_norm for k in ["device", "단말", "list", "spec"]):
            spec_candidates.append(s)
    st.write("📱 단말 스펙 시트 후보:", ", ".join(spec_candidates) if spec_candidates else "(없음)")

    spec_sheets_selected = st.multiselect(
        "Device List / Spec 시트를 선택하세요 (없으면 생략 가능).",
        options=sheet_names,
        default=spec_candidates,
    )

# ==============================
# 3) 테스트 시트에서 Fail + 코멘트 추출
# ==============================
with step_status("테스트 시트에서 Fail + 코멘트 추출"):
    try:
        wb_comm = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
        wb_val  = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

        available = set(wb_comm.sheetnames) & set(wb_val.sheetnames)
        valid_sheets = [s for s in test_sheets_selected if s in available]
        if not valid_sheets:
            st.error(f"선택한 시트를 찾을 수 없습니다. 사용 가능: {sorted(list(available))}")
            st.stop()

        # 스레드 댓글까지 읽어서 보강
        threaded_map = load_threaded_comments_map_from_bytes(data)

        df_issue = extract_comments_as_dataframe_dual(
            wb_comm, wb_val, valid_sheets, threaded_map=threaded_map
        )
        # 디버그용 확장은 필요할 때만 열어보면 되므로, 화면에는 최종 병합 뷰(df_final)만 보여준다.
        # diag_dump("추출된 Fail+코멘트 전체", df_issue)

        if df_issue.empty:
            st.warning("❌ Fail+코멘트 항목이 없습니다(메모/댓글 미검출).")
            st.info("엑셀에서 해당 셀에 실제 코멘트가 존재하는지(새 댓글/메모), 보호/숨김 상태가 아닌지 확인해 주세요.")
            st.stop()
    except Exception as e:
        st.error(f"코멘트 추출 중 오류: {str(e)}")
        st.stop()

# ==============================
# 4) 스펙 병합 (모델명 정규화 후 Join—헤더 자동탐지 + 부분일치 백업)
# ==============================
df_final = df_issue.copy()
match_rate = 0.0

# df_final에 Device(Model) 컬럼이 없으면, 원본 헤더 후보에서 복사해 생성
if "Device(Model)" not in df_final.columns:
    for cand in ["Device", "device", "Model", "MODEL", "단말명", "모델명"]:
        if cand in df_final.columns:
            df_final["Device(Model)"] = df_final[cand]
            break

df_spec_all = pd.DataFrame()
spec_match_info = None
df_spec_mismatch_sample = pd.DataFrame()

if spec_sheets_selected:
    # ⏱ 스펙 병합 단계 타이틀 없이 내부에서만 병합 수행
    # ---------- 공통 유틸 ----------
    def _norm_hdr(s: str) -> str:
        s = unicodedata.normalize("NFKC", str(s))
        s = re.sub(r"[\s\-\_/()\[\]{}:+·∙•]", "", s).lower()
        return s

    def find_header_row_for_spec(xls, sheet, max_scan_rows=20):
        """스펙 시트에서 헤더 행(모델 관련 키워드가 포함된 행)을 위에서부터 탐색"""
        probe = pd.read_excel(xls, sheet_name=sheet, header=None, engine="openpyxl")
        header_keywords = [r"^model$", r"^device$", r"^제품명$", r"^모델$", r"^모델명$", r"^기종$", r"^단말$", r"^단말명$"]
        for i in range(min(max_scan_rows, len(probe))):
            row = probe.iloc[i].astype(str).fillna("")
            hits = 0
            for cell in row:
                c = _norm_hdr(cell)
                if any(re.search(pat, c) for pat in header_keywords):
                    hits += 1
            if hits >= 1:
                return i
        return 0

    def standardize_spec_columns(df: pd.DataFrame) -> pd.DataFrame:
        orig = list(df.columns)
        norm = [_norm_hdr(c) for c in orig]
        col_map = {}
        synonyms = {
            r"^(model|device|제품명|제품|모델명|모델|단말|단말명|기종)$": "Model",
            r"^(maker|manufacturer|brand|oem|제조사|벤더)$": "제조사",
            r"^(gpu|gpu명|gpumodel|graphics|그래픽|그래픽칩|그래픽스|그래픽프로세서)$": "GPU",
            r"^(chipset|soc|ap|cpu|processor)$": "Chipset",
            r"^(ram|메모리)$": "RAM",
            r"^(os|osversion|android|ios|펌웨어|소프트웨어버전|운영체제|os버전)$": "OS",
            r"^(rank|rating|ratinggrade|등급)$": "Rank",
            # 해상도/디스플레이 해상도
            r"^(resolution|해상도|display|displayresolution|resolutiondisplay)$": "Resolution",
        }
        for n, o in zip(norm, orig):
            mapped = None
            for pat, std in synonyms.items():
                if re.search(pat, n):
                    mapped = std
                    break
            col_map[o] = mapped or o
        return df.rename(columns=col_map)

    # ---------- 스펙 시트 적재 ----------
    frames = []
    for sname in spec_sheets_selected:
        try:
            hdr = find_header_row_for_spec(xls, sname)
            dfp = pd.read_excel(xls, sheet_name=sname, header=hdr, engine="openpyxl")
        except Exception:
            continue
        dfp = standardize_spec_columns(dfp)

        # 필수: Model 열
        model_col = "Model" if "Model" in dfp.columns else None
        if not model_col:
            for c in dfp.columns:
                if re.search(r"(model|device|제품명|모델|기종|단말)", _norm_hdr(c)):
                    model_col = c; break
        if not model_col:
            continue

        # 정규화 키 생성
        from qa_patch_module import normalize_model_name_strict
        dfp["model_norm"] = dfp[model_col].apply(normalize_model_name_strict)

        # 보조 키(색상·용량 제거 전 원문도 보관)
        dfp["model_raw"] = dfp[model_col].astype(str)

        # 유지 컬럼
        keep = ["model_norm", "model_raw"] + [
            c for c in ["GPU","제조사","Chipset","RAM","OS","Rank","Model","CPU","Resolution"]
            if c in dfp.columns
        ]
        frames.append(dfp[keep])

    if not frames:
        st.warning("선택한 스펙 시트에서 유효한 헤더/모델 열을 찾지 못했습니다. (헤더 위치/열 이름 확인)")
    else:
        df_spec_all = pd.concat(frames, ignore_index=True).drop_duplicates("model_norm", keep="first")

        # ---------- 이슈쪽 모델 정규화 ----------
        from qa_patch_module import normalize_model_name_strict
        df_final["model_norm"] = df_final["Device(Model)"].apply(normalize_model_name_strict)

        # 1차: model_norm으로 정석 병합
        df_final = pd.merge(df_final, df_spec_all, on="model_norm", how="left", suffixes=("","_spec"))

        # Chipset 보정
        if "Chipset" not in df_final.columns and "CPU" in df_final.columns:
            df_final["Chipset"] = df_final["CPU"]

        # 접미사 정리
        for col in ["GPU","제조사","Chipset","RAM","OS","Rank","Model"]:
            cx, cy = f"{col}", f"{col}_spec"
            if cx in df_final.columns and cy in df_final.columns:
                df_final[col] = df_final[cx].where(df_final[cx].notna() & (df_final[cx]!=""), df_final[cy])
                df_final.drop(columns=[cy], inplace=True, errors="ignore")
            elif cy in df_final.columns:
                df_final.rename(columns={cy: col}, inplace=True)

        # ---------- 2차: 부분일치(contains) 백업 매칭 ----------
if "GPU" in df_final.columns:
    mask_need = (
        df_final["GPU"].isna()
        | (df_final["GPU"].astype(str).str.strip() == "")
    ) & (df_final["Device(Model)"].astype(str).str.len() > 0)

    if mask_need.any() and not df_spec_all.empty:
        base_cols = ["model_raw", "GPU", "Chipset", "OS", "Rank"]
        existing_cols = [c for c in base_cols if c in df_spec_all.columns]

        if "model_raw" not in existing_cols:
            st.info("⚠ df_spec_all에 model_raw 컬럼이 없어 부분 매칭을 생략합니다.")
        else:
            spec_index = (
                df_spec_all[existing_cols]
                .dropna(subset=["model_raw"])
                .reset_index(drop=True)
            )

            for idx in df_final[mask_need].index.tolist():
                key = str(df_final.at[idx, "Device(Model)"])
                hit = spec_index[spec_index["model_raw"].astype(str).str.contains(key, case=False, na=False)]
                if not hit.empty:
                    h0 = hit.iloc[0].to_dict()
                    for col in ["GPU", "Chipset", "OS", "Rank"]:
                        if col in h0 and pd.isna(df_final.at[idx, col]):
                            df_final.at[idx, col] = h0.get(col, "")

# 스펙 매칭 요약 정보(나중에 별도 UI 섹션에서 사용)
if "GPU" in df_final.columns:
    matched = int(df_final["GPU"].fillna("").astype(str).str.strip().ne("").sum())
    match_rate = round(matched / max(1, len(df_final)) * 100, 1)
    spec_match_info = {
        "matched": matched,
        "total": int(len(df_final)),
        "match_rate": match_rate,
    }
    df_spec_mismatch_sample = df_final[
        df_final["GPU"].fillna("").astype(str).str.strip() == ""
    ][["Device(Model)", "GPU", "Chipset", "OS", "Rank"]].head(20)

# ==============================
# 6) 자가진단 (내부 로직만 수행, 웹 출력은 최소화)
# ==============================
diag = self_check(df_final)
if not diag["row_ok"]:
    st.error("❌ 유효한 데이터 없음. 중단.")
    st.stop()

# ==============================
# 7) 코멘트 정규화/태깅 (계산만 먼저 수행)
# ==============================

def _jamo_norm(s: str) -> str:
    if s is None: return ""
    t = unicodedata.normalize("NFKC", str(s))
    t = re.sub(r"[^0-9a-zA-Z가-힣\s\-_+/.:]", " ", t)
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t

ISSUE_TAG_PATTERNS = [
    ("punch_hole", r"(펀치홀|punch[\s\-]?hole|hole[-\s]?camera)"),
    ("notch", r"(노치|notch)"),
    ("rotation", r"(회전|가로전환|세로전환|landscape|portrait|rotate)"),
    ("aspect_ratio", r"(화면비|비율|aspect\s?ratio)"),
    ("resolution", r"(해상도|resolution)"),
    ("cutout", r"(컷아웃|cutout)"),
    ("install", r"(설치\s?불가|설치오류|install\s?fail|패키지\s?오류|apk\s?설치)"),
    ("permission", r"(권한|permission)"),
    ("login", r"(로그인|login|oauth|인증|auth)"),
    ("storage", r"(저장공간|storage|sd\s?card|권한\s?거부)"),
    ("input_lag", r"(입력\s?지연|터치\s?지연|ui\s?지연|input\s?lag)"),
    ("crash", r"(크래시|crash|강제종료|프로세스\s?종료)"),
    ("freeze", r"(멈춤|버벅임|프리즈|freeze)"),
    ("network", r"(네트워크|network|핑|ping|latency|disconnect)"),
    ("render", r"(렌더링|render|그림자|텍스처|texture|shader)"),
    ("ui_scaling", r"(작게\s?보임|축소|缩小|small ui|스케일링|scaling|해상도\s?고정|1080p)"),
    ("ui_margin", r"(좌측\s?여백|여백\s?발생|margin|padding)"),
    ("option_graphics", r"(그래픽\s?옵션|옵션\s?화면|settings|option)"),
    ("frame_cap", r"(프레임\s?설정|fps\s?제한|60fps|120fps)"),
    ("audio", r"(소리|오디오|audio|무음|볼륨)"),
    ("camera", r"(카메라|camera)"),
    ("thermal", r"(써멀|발열|thermal|throttl)"),
    ("fps", r"(프레임|fps)"),
]

def tag_issue_comment(comment: str) -> list:
    s = _jamo_norm(comment)
    tags = []
    for tag, pat in ISSUE_TAG_PATTERNS:
        if re.search(pat, s, re.I):
            tags.append(tag)
    return list(dict.fromkeys(tags))

# comment_text / issue_tags 생성
if "comment_text" not in df_final.columns:
    if "comment_cell" in df_final.columns:
        df_final["comment_text"] = df_final["comment_cell"].fillna("").astype(str)
    else:
        df_final["comment_text"] = ""

def _strip_excel_thread_prefix(s: str) -> str:
    if s is None:
        return ""
    text = str(s)
    m = re.search(r"댓글\s*:\s*", text)
    if m:
        return text[m.end():].strip()
    return text.strip()

df_final["comment_text"] = df_final["comment_text"].astype(str).apply(_strip_excel_thread_prefix)

df_final["comment_norm"] = (
    df_final["comment_text"]
    .fillna("")
    .astype(str)
    .apply(_jamo_norm)
)
df_final["issue_tags"] = (
    df_final["comment_text"]
    .fillna("")
    .astype(str)
    .apply(tag_issue_comment)
)

for col in ["Device(Model)", "GPU", "Chipset", "OS"]:
    if col not in df_final.columns:
        df_final[col] = ""

# 스펙 병합 + 코멘트/태깅이 반영된 최종 Fail 뷰만 출력
# issue_tags를 마지막 컬럼으로 함께 노출
cols_show = [
    c for c in [
        "Sheet","Device(Model)","GPU","Chipset","RAM","OS","Resolution","Rank","Checklist","comment_text","issue_tags"
    ] if c in df_final.columns
]
st.write("### 테스트 시트에서 Fail + 코멘트 + 스펙 (최종)")
st.dataframe(df_final[cols_show])

# ==============================
# 7-1) 스펙 병합 요약 UI
# ==============================
with step_status("스펙 병합 요약"):
    if spec_match_info is not None:
        st.success(
            f"스펙 매칭 결과: GPU 채움 {spec_match_info['matched']} / "
            f"{spec_match_info['total']} 건 ({spec_match_info['match_rate']}%)"
        )
        if not df_spec_mismatch_sample.empty:
            diag_dump(
                "스펙 병합 미매칭 샘플(상위 20)",
                df_spec_mismatch_sample,
            )
    else:
        st.info("스펙 시트가 선택되지 않았거나 매칭 정보가 없습니다.")

# ==============================
# 7-2) 코멘트 정규화 / 태깅 상태 UI
# ==============================
with step_status("코멘트 정규화 / 태깅"):
    st.write("comment_text / issue_tags 정규화 및 태깅이 완료되었습니다.")

# ==============================
# 8) 군집 산출
# ==============================
with step_status("군집(Cluster) 통계 산출"):
    if "Chipset" not in df_final.columns and "CPU" in df_final.columns:
        df_final["Chipset"] = df_final["CPU"]

    if "GPU" not in df_final.columns:
        df_final["GPU"] = ""

    # 디버그용 클러스터 통계 (웹에서 확인 가능)
    try:
        # ----------------------------
        # 1) 이슈 × GPU × Chipset × 해상도 군집(issue_hw)
        # ----------------------------
        needed = {"issue_tags", "GPU", "Chipset", "Resolution", "Device(Model)"}
        if needed.issubset(df_final.columns):
            tmp = df_final.copy()
            tmp["GPU"] = tmp["GPU"].fillna("").astype(str).str.strip()
            tmp["Chipset"] = tmp["Chipset"].fillna("").astype(str).str.strip()
            tmp["Resolution"] = tmp["Resolution"].fillna("").astype(str).str.strip()

            # 스펙 정보가 비어 있으면 제외
            tmp = tmp[
                (tmp["GPU"] != "")
                & (tmp["Chipset"] != "")
                & (tmp["Resolution"] != "")
            ]

            if not tmp.empty and "issue_tags" in tmp.columns:
                ex = tmp.explode("issue_tags")
                ex["issue_tags"] = ex["issue_tags"].fillna("").astype(str).str.strip()
                ex = ex[ex["issue_tags"] != ""]

                if not ex.empty:
                    df_issue_hw = (
                        ex.groupby(["issue_tags", "GPU", "Chipset", "Resolution"])[
                            "Device(Model)"
                        ]
                        .agg(
                            fail_device_count=lambda s: s.dropna().nunique(),
                            repr_models=lambda s: ", ".join(
                                sorted(set(map(str, s.dropna())))[:5]
                            ),
                        )
                        .reset_index()
                        .sort_values(
                            "fail_device_count", ascending=False
                        )
                    )
                    diag_dump(
                        "이슈 × GPU × Chipset × 해상도 군집(issue_hw)",
                        df_issue_hw,
                    )

        # ----------------------------
        # 2) 단일 축 기준 통계 + issue_tag × 축 별 집계
        #    - GPU / Chipset / Resolution 각각에 대해
        # ----------------------------
        cols = set(df_final.columns)

        # (a) GPU 기준
        if {"GPU", "Device(Model)"}.issubset(cols):
            df_gpu = df_final.assign(GPU=df_final["GPU"].fillna("").astype(str).str.strip())
            df_gpu = df_gpu[df_gpu["GPU"] != ""]

            # GPU 단독 통계
            df_cluster_gpu = (
                df_gpu.groupby("GPU")["Device(Model)"]
                .nunique()
                .reset_index(name="fail_device_count")
                .sort_values("fail_device_count", ascending=False)
            )
            if not df_cluster_gpu.empty:
                diag_dump("클러스터 통계 - GPU별 Fail 단말 수", df_cluster_gpu)

            # issue_tag × GPU 집계
            if "issue_tags" in df_gpu.columns:
                ex = df_gpu.explode("issue_tags")
                ex["issue_tags"] = ex["issue_tags"].fillna("").astype(str).str.strip()
                ex = ex[ex["issue_tags"] != ""]
                if not ex.empty:
                    df_by_gpu = (
                        ex.groupby(["issue_tags", "GPU"])["Device(Model)"]
                        .nunique()
                        .reset_index(name="fail_device_count")
                        .sort_values("fail_device_count", ascending=False)
                    )
                    diag_dump("이슈×GPU 집계(by_gpu_cluster)", df_by_gpu)

        # (b) Chipset(CPU) 기준
        if {"Chipset", "Device(Model)"}.issubset(cols):
            df_chip = df_final.assign(Chipset=df_final["Chipset"].fillna("").astype(str).str.strip())
            df_chip = df_chip[df_chip["Chipset"] != ""]

            df_cluster_chip = (
                df_chip.groupby("Chipset")["Device(Model)"]
                .nunique()
                .reset_index(name="fail_device_count")
                .sort_values("fail_device_count", ascending=False)
            )
            if not df_cluster_chip.empty:
                diag_dump("클러스터 통계 - Chipset별 Fail 단말 수", df_cluster_chip)

            if "issue_tags" in df_chip.columns:
                ex = df_chip.explode("issue_tags")
                ex["issue_tags"] = ex["issue_tags"].fillna("").astype(str).str.strip()
                ex = ex[ex["issue_tags"] != ""]
                if not ex.empty:
                    df_by_chip = (
                        ex.groupby(["issue_tags", "Chipset"])["Device(Model)"]
                        .nunique()
                        .reset_index(name="fail_device_count")
                        .sort_values("fail_device_count", ascending=False)
                    )
                    diag_dump("이슈×Chipset 집계(by_chipset_cluster)", df_by_chip)

        # (c) 해상도 기준
        if {"Resolution", "Device(Model)"}.issubset(cols):
            df_res = df_final.assign(
                Resolution=df_final["Resolution"].fillna("").astype(str).str.strip()
            )
            df_res = df_res[df_res["Resolution"] != ""]

            df_cluster_res = (
                df_res.groupby("Resolution")["Device(Model)"]
                .nunique()
                .reset_index(name="fail_device_count")
                .sort_values("fail_device_count", ascending=False)
            )
            if not df_cluster_res.empty:
                diag_dump("클러스터 통계 - 해상도별 Fail 단말 수", df_cluster_res)

            if "issue_tags" in df_res.columns:
                ex = df_res.explode("issue_tags")
                ex["issue_tags"] = ex["issue_tags"].fillna("").astype(str).str.strip()
                ex = ex[ex["issue_tags"] != ""]
                if not ex.empty:
                    df_by_res = (
                        ex.groupby(["issue_tags", "Resolution"])["Device(Model)"]
                        .nunique()
                        .reset_index(name="fail_device_count")
                        .sort_values("fail_device_count", ascending=False)
                    )
                    diag_dump("이슈×해상도 집계(by_resolution_cluster)", df_by_res)
    except Exception as e:
        diag_dump("클러스터 통계 계산 오류", str(e))

# ==============================
# 9) 메트릭 계산 (내부 계산만 수행, 웹 출력 없음)
# ==============================
metrics = {}

total_rows = len(df_final)
metrics["total_rows"] = total_rows
metrics["total_fail_issues"] = total_rows

if "issue_tags" in df_final.columns:
    exploded = df_final.explode("issue_tags")
    vc = exploded["issue_tags"].value_counts().reset_index(name="count")
    vc = vc.rename(columns={"issue_tags": "value"})
    tag_counts = vc.to_dict(orient="records")
else:
    tag_counts = []
metrics["by_issue_tag"] = tag_counts

# 이슈 × GPU × Chipset × 해상도 군집(issue_hw) – 메트릭용
clusters_issue_hw = []
needed_cols = {"issue_tags", "GPU", "Chipset", "Resolution", "Device(Model)"}
if needed_cols.issubset(df_final.columns):
    tmp = df_final.copy()
    tmp["GPU"] = tmp["GPU"].fillna("").astype(str).str.strip()
    tmp["Chipset"] = tmp["Chipset"].fillna("").astype(str).str.strip()
    tmp["Resolution"] = tmp["Resolution"].fillna("").astype(str).str.strip()

    tmp = tmp[
        (tmp["GPU"] != "")
        & (tmp["Chipset"] != "")
        & (tmp["Resolution"] != "")
    ]

    if not tmp.empty:
        ex = tmp.explode("issue_tags")
        ex["issue_tags"] = ex["issue_tags"].fillna("").astype(str).str.strip()
        ex = ex[ex["issue_tags"] != ""]

        if not ex.empty:
            grp = (
                ex.groupby(["issue_tags", "GPU", "Chipset", "Resolution"])[
                    "Device(Model)"
                ]
                .agg(lambda s: sorted(set(map(str, s.dropna()))))
                .reset_index()
            )

            for _, r in grp.iterrows():
                models = r["Device(Model)"]
                count = len(models)
                clusters_issue_hw.append(
                    {
                        "feature_tag": r["issue_tags"],
                        "hw_type": "gpu+chipset",
                        "hw_value": f"{r['GPU']} / {r['Chipset']}",
                        "resolution_group": r["Resolution"],
                        "count": count,
                        "repr_models": models[:5],
                    }
                )

metrics["clusters"] = {"issue_hw": clusters_issue_hw}
metrics["clusters_feature_detailed"] = []

meta = {
    "build_version": "KP 4.2.0",
    "scope": "Android / iOS 호환성 검증",
    "metrics": metrics,
}

# ==============================
# 10) LLM 호출 (gpt-5.1, JSON 객체 강제)
# ==============================
sp = build_system_prompt()
up = build_user_prompt(df_final, meta)

st.write("### 🤖 OpenAI 호출 (필요 시 모델만 교체: gpt-5.1)")
with st.spinner("GPT가 리포트를 작성 중입니다..."):
    max_retries, wait = 3, 20
    result, last_error = None, None
    for attempt in range(max_retries):
        try:
            resp = client.chat.completions.create(
                model="gpt-5.1",          # 품질 우선
                temperature=0.1,
                top_p=0.9,
                messages=[{"role":"system","content":sp},{"role":"user","content":up}],
            )
            raw = resp.choices[0].message.content
            result = parse_llm_json(raw)
            result["meta"] = meta
            diag_dump("LLM 원문(요약)", raw[:3000])
            break
        except Exception as e:
            last_error = e
            if "429" in str(e) or "rate_limit_exceeded" in str(e).lower():
                if attempt < max_retries-1:
                    st.warning(f"429 감지, 재시도 {attempt+1}/{max_retries}")
                    time.sleep(wait); wait *= 2
                    continue
            st.error(f"OpenAI 호출 실패: {e}")
            st.stop()
    if result is None:
        st.error(f"OpenAI 최종 실패: {last_error}")
        st.stop()

# ==============================
# 11) 리포트 생성 (4개 시트)
# ==============================
# 11) 리포트 생성
try:
    # df_spec_all이 스펙 병합 시점에 만들어져 있다면 단말 전체 리스트로 사용
    try:
        df_devices_all = df_spec_all.copy()
    except NameError:
        # 스펙 시트가 없을 경우 최소한 df_final 기반으로라도 생성
        df_devices_all = df_final.copy()

    # Platform 컬럼 보정 (없으면 AOS로 기본값)
    if "Platform" not in df_devices_all.columns:
        if "OS" in df_devices_all.columns:
            df_devices_all["Platform"] = df_devices_all["OS"].apply(
                lambda x: "iOS" if str(x).lower().startswith("ios") else "AOS"
            )
        else:
            df_devices_all["Platform"] = "AOS"

    # Device(Model) 컬럼 보정
    if "Device(Model)" not in df_devices_all.columns:
        for cand in ["Device", "device", "Model", "MODEL", "단말명", "모델명"]:
            if cand in df_devices_all.columns:
                df_devices_all["Device(Model)"] = df_devices_all[cand]
                break

    output_path = "QA_Report_4sheets.xlsx"

    write_excel_report(
        result=result,
        df_final=df_final,
        df_devices_all=df_devices_all,
        path=output_path,
    )

    st.success("✅ 4개 시트 포함 리포트 생성 완료")
    with open(output_path, "rb") as f:
        st.download_button("📊 Excel 리포트 다운로드", f.read(), file_name=output_path)
except Exception as e:
    st.error(f"리포트 생성 오류: {e}")

