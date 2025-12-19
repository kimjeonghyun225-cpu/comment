# -*- coding: utf-8 -*-
"""
QA 통합 패치 모듈 (최종)
- 테스트 시트 자동 감지
- Fail 셀 + 코멘트 추출 (라벨행→Fail열 세로추출, 병합셀 보정, 수식/스레드댓글 대응)
- 셀 코멘트 + 비고/Notes 병합
- LLM JSON 강제(A~E 스펙, issues[] 제한 없음)
- Excel 리포트(Executive_Summary / Summary / Issues / Device_Risks / Evidence_Sample / Cluster_*)
- 모듈 자가진단(self_check)
"""
import io
import re
import json
import zipfile
import unicodedata
import xml.etree.ElementTree as ET
from typing import List, Dict, Any, Optional

import pandas as pd
import openpyxl

# 이슈 태그 → 한국어 Fail Category 라벨 매핑
TAG_LABEL_KO = {
    "punch_hole":      "펀치홀 영역 문제",
    "notch":           "노치 영역 문제",
    "ui_clipped":      "UI 잘림",
    "ui_overlap":      "UI 겹침",
    "safearea":        "Safe Area 처리 문제",
    "rotation":        "회전/가로·세로 모드 문제",
    "ratio":           "화면 비율 문제",
    "resolution":      "해상도 처리 문제",
    "cutout":          "컷아웃 처리 문제",
    "install":         "설치 실패/설치 오류",
    "permission":      "권한 설정 문제",
    "login":           "로그인/인증 문제",
    "storage":         "저장공간/SD카드 문제",
    "input_lag":       "입력 지연/터치 랙",
    "crash":           "크래시/강제 종료",
    "freeze":          "멈춤/프리즈",
    "network":         "네트워크/핑 문제",
    "render":          "렌더링/텍스처 문제",
    "ui_scaling":      "UI 축소/스케일링 문제",
    "ui_margin":       "UI 여백/레이아웃 문제",
    "option_graphics": "그래픽 옵션 화면 문제",
    "frame_cap":       "프레임 설정 제한",
    "audio":           "오디오/소리 문제",
    "camera":          "카메라 동작 문제",
    "thermal":         "발열/써멀 스로틀링",
    "fps":             "프레임 변동/프레임 드랍",
}

# ==============================
# 공통 유틸
# ==============================
def _norm_for_header(s: str) -> str:
    s = unicodedata.normalize("NFKC", str(s))
    s = re.sub(r"[\s\-\_/()\[\]{}:+·∙•]", "", s)
    return s.lower().strip()

def _nkfc(s: Any) -> str:
    if pd.isna(s): return ""
    s = str(s)
    s = re.sub(r"\s+", " ", s)
    return unicodedata.normalize("NFKC", s).strip()

def normalize_model_name_strict(model: str) -> str:
    """
    모델명 정규화: 괄호/용량/색상 제거, 공백/기호 정리, 소문자 통일.
    동일 기종 매칭을 위한 키 생성 용도.
    """
    if model is None:
        return ""
    s = unicodedata.normalize("NFKC", str(model))
    s = re.sub(r"\(.*?\)", " ", s)
    s = re.sub(r"\b(4g|5g|5g uw|lte|dual sim)\b", " ", s, flags=re.I)
    s = re.sub(r"\b(128gb|256gb|512gb|1tb)\b", " ", s, flags=re.I)
    s = re.sub(r"\b(black|white|blue|green|red|pink|silver|gold|purple|violet)\b", " ", s, flags=re.I)
    s = re.sub(r"[\+\-_/]", " ", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

# ==============================
# 병합셀 안전 읽기 + 라벨행 탐지
# ==============================
def read_merged(ws, r: int, c: int):
    """병합 셀을 고려해서 (r,c)의 실제 값을 안전하게 읽는다."""
    for rng in ws.merged_cells.ranges:
        if (rng.min_row <= r <= rng.max_row) and (rng.min_col <= c <= rng.max_col):
            return ws.cell(row=rng.min_row, column=rng.min_col).value
    return ws.cell(row=r, column=c).value


def find_row_by_labels(ws, labels, search_rows: int = 30, search_cols: int = 80) -> int:
    """
    주어진 라벨 텍스트들 중 하나가 포함된 셀을 상단 영역에서 찾고,
    해당 셀이 위치한 행 번호를 반환한다. (없으면 0)
    """
    labels_norm = {_norm_for_header(x) for x in labels}
    max_r = min(search_rows, ws.max_row)
    max_c = min(search_cols, ws.max_column)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = read_merged(ws, r, c)
            if v and _norm_for_header(v) in labels_norm:
                return r
    return 0

# ==============================
# 테스트 시트 후보 탐지
# ==============================
def find_test_sheet_candidates(names: List[str]) -> List[str]:
    cands = set()
    for n in names:
        norm = re.sub(r"[\s_\-]+", "", n.lower())
        if any(k in norm for k in ["compatibility", "compattest", "test", "qa", "호환성", "테스트", "tc_"]):
            cands.add(n)
    if cands:
        return sorted(cands)
    # 후보가 없으면 전체 반환
    return names

# ==============================
# 스레드 댓글 파싱
# ==============================
def _sanitize_excel_comment(text: str) -> str:
    """Excel 스레드 댓글/안내문 머리말을 제거하고 사용자 본문만 남긴다."""
    s = str(text or "")
    lines = [ln for ln in s.splitlines()]
    if lines:
        if re.match(r"^\s*\[?\s*스레드\s*댓글[^\]]*\]?\s*$", lines[0], flags=re.I):
            lines = lines[1:]
        elif re.match(r"^\s*this\s+cell\s+contains\s+a\s+threaded\s+comment", lines[0], flags=re.I):
            lines = lines[1:]
    s = "\n".join(lines)
    s = s.replace("https://go.microsoft.com/fwlink/?linkid=870924.", "")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s).strip()
    return s

def load_threaded_comments_map_from_bytes(xlsx_bytes: bytes) -> dict:
    """
    .xlsx 바이트에서 스레드 댓글을 파싱해 {(sheet_name, cell_ref): "text"} 사전으로 반환.
    여러 댓글이 한 셀에 붙은 경우 ' | '로 이어 붙임.
    """
    mapping = {}  # key: (sheet_name, ref), value: list[str]

    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as zf:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
            sheets = list(wb.sheetnames)
        except Exception:
            sheets = []

        for sheet_name in sheets:
            sheet_part = None
            for p in zf.namelist():
                if p.startswith("xl/worksheets/") and p.endswith(".xml"):
                    if sheet_name in p:
                        sheet_part = p
                        break
            if not sheet_part:
                continue

            # 같은 폴더 안의 _rels에 threadedComment 참조가 있을 수 있음
            base = "/".join(sheet_part.split("/")[:-1])
            candidates = []
            for p in zf.namelist():
                if p.startswith(base) and "threaded" in p and p.endswith(".xml"):
                    candidates.append(p)
            for tc_path in candidates:
                try:
                    tc_xml = ET.fromstring(zf.read(tc_path))
                    for tcm in tc_xml.iter():
                        if tcm.tag.endswith("threadedComment"):
                            cell_ref = tcm.get("ref")
                            if not cell_ref:
                                continue
                            texts = []
                            for node in tcm.iter():
                                if node.text and node.tag.endswith(("t", "text")):
                                    texts.append(str(node.text))
                            txt = _sanitize_excel_comment(" ".join([x.strip() for x in texts if x and x.strip()]))
                            if not txt:
                                continue
                            key = (sheet_name, cell_ref)
                            mapping.setdefault(key, []).append(txt)
                except Exception:
                    continue

    merged = {}
    for k, v in mapping.items():
        merged[k] = " | ".join([x for x in v if x])
    return merged

# ==============================
# Fail + 코멘트 추출 (라벨행→Fail열 세로추출)
# ==============================
def extract_comments_as_dataframe_dual(
    wb_comments: openpyxl.Workbook,
    wb_values: openpyxl.Workbook,
    target_sheet_names: List[str],
    threaded_map: Optional[Dict[Any, Any]] = None,
) -> pd.DataFrame:
    """
    Fail 셀과 코멘트를 찾고, 해당 Fail "열(column)"에서
    Model/GPU/Chipset/RAM/OS/Rank 값을 라벨행 기준으로 수직 추출.
    병합셀도 보정하여 정확도 확보.
    threaded_map: {(sheet_name, "A1"): "threaded text"} 보강에 사용.
    """
    if threaded_map is None:
        threaded_map = {}

    extracted = []
    for sheet_name in target_sheet_names:
        if sheet_name not in wb_comments.sheetnames or sheet_name not in wb_values.sheetnames:
            continue
        ws_c = wb_comments[sheet_name]
        ws_v = wb_values[sheet_name]

        # 각 스펙 라벨이 있는 "행 번호"를 찾는다.
        header_rows = {
            "Model":   find_row_by_labels(ws_c, ["Model","Device","제품명","제품","모델명","모델","단말","단말명","기종"]),
            "GPU":     find_row_by_labels(ws_c, ["GPU","그래픽","그래픽칩","그래픽스","그래픽프로세서"]),
            "Chipset": find_row_by_labels(ws_c, ["Chipset","SoC","AP","CPU","Processor","칩셋"]),
            "RAM":     find_row_by_labels(ws_c, ["RAM","메모리"]),
            "OS":      find_row_by_labels(ws_c, ["OS Version","Android","iOS","OS","펌웨어","소프트웨어버전"]),
            "Rank":    find_row_by_labels(ws_c, ["Rating Grade?","Rank","등급"]),
        }

        def _get_val(rr: int, cc: int) -> str:
            """값 워크북에서 병합·수식 등을 보정하여 안전하게 값 읽기."""
            if rr <= 0:
                return ""
            v = read_merged(ws_v, rr, cc)
            # 숫자는 문자열로 정규화
            if isinstance(v, (int, float)):
                v = str(int(v)) if isinstance(v, float) and v.is_integer() else str(v)
            # 수식 셀은 값 무시
            if isinstance(v, str) and v.startswith("="):
                return ""
            return (str(v) if v is not None else "").strip()

        # 코멘트 워크시트를 모두 훑으면서 'Fail' 셀을 찾는다.
        fail_total = 0
        fail_with_comment = 0
        for row in ws_c.iter_rows():
            for cell in row:
                val = cell.value
                # 공백 포함 'Fail' 정확 매칭 (대소문자 무시)
                if isinstance(val, str) and re.fullmatch(r"\s*fail\s*", val, flags=re.I):
                    r, c = cell.row, cell.column

                    # 1) 구형 메모(comment)
                    raw_comment = (getattr(cell, "comment", None).text
                                   if getattr(cell, "comment", None) else "")
                    ctext = _sanitize_excel_comment(raw_comment)

                    # 2) threadedComments 보강
                    if not ctext:
                        key = (sheet_name, cell.coordinate)  # (시트명, "A1")
                        if key in threaded_map:
                            ctext = _sanitize_excel_comment(threaded_map[key])

                    fail_total += 1

                    # 3) 여전히 코멘트가 없으면 이 Fail은 스킵
                    if not ctext:
                        continue
                    fail_with_comment += 1

                    # Fail이 찍힌 그 "열"에서, 각 스펙 라벨 행의 값을 읽어온다.
                    device_info = {key: _get_val(rr, c) for key, rr in header_rows.items()}

                    extracted.append({
                        "Sheet": ws_c.title,
                        "Device(Model)": device_info.get("Model", ""),
                        "GPU":     device_info.get("GPU", ""),
                        "Chipset": device_info.get("Chipset", ""),
                        "RAM":     device_info.get("RAM", ""),
                        "OS":      device_info.get("OS", ""),
                        "Rank":    device_info.get("Rank", ""),
                        # Checklist는 시트명으로 두고, 나중에 Notes 병합에서 _nkfc로 맞춰 씀
                        "Checklist": ws_c.title,
                        "comment_cell": ctext,
                        "Comment(Text)": "",
                    })

        # 시트별 디버그 로그
        print(
            f"[extract_comments_as_dataframe_dual] sheet={sheet_name} "
            f"Fail셀={fail_total}, 코멘트있는 Fail={fail_with_comment}, "
            f"누적추출행수={len(extracted)}"
        )

    if not extracted:
        return pd.DataFrame(columns=[
            "Sheet","Device(Model)","GPU","Chipset","RAM","OS","Rank",
            "Checklist","comment_cell","Comment(Text)"
        ])
    return pd.DataFrame(extracted)

# ==============================
# 비고/Notes 병합 (현재는 사용 안 함 – 패스스루)
# ==============================
def enrich_with_column_comments(
    xls_bytes: io.BytesIO,
    sheet_name: str,
    df_issue: pd.DataFrame
) -> pd.DataFrame:
    """
    이전 버전에서는 시트 내 '비고/Notes' 열을 코멘트에 병합했지만,
    현재는 셀 코멘트(comment_cell)만 사용하므로 그대로 반환한다.
    """
    return df_issue

# ==============================
# self_check (간단 진단)
# ==============================
def self_check(df_final: pd.DataFrame) -> Dict[str, Any]:
    row_ok = len(df_final) > 0
    has_device = "Device(Model)" in df_final.columns
    # comment_text는 코멘트 정규화/태깅 단계에서 생성되므로,
    # 그 이전 단계에서는 comment_cell 존재 여부도 함께 인정해준다.
    has_comment = ("comment_text" in df_final.columns) or ("comment_cell" in df_final.columns)
    return {
        "row_ok": bool(row_ok),
        "has_device": bool(has_device),
        "has_comment": bool(has_comment),
    }

# ==============================
# LLM JSON 파서
# ==============================
def parse_llm_json(raw: str) -> Dict[str, Any]:
    raw = raw.strip()
    try:
        return json.loads(raw)
    except Exception:
        try:
            m = re.search(r"\{.*\}", raw, flags=re.S)
            if m:
                return json.loads(m.group(0))
        except Exception:
            pass
    return {"raw": raw}

# ==============================
# 프롬프트 빌더
# ==============================
def build_system_prompt() -> str:
    return (
        "당신은 게임 호환성 및 성능 QA 리포트를 작성하는 전문가입니다. "
        "입력으로 단말별 Fail 코멘트, 스펙(GPU/Chipset/OS/Rank) 정보가 주어지며, "
        "이를 기반으로 **데이터에 근거한 내용만** 요약/분석해야 합니다. "
        "추측, 가정, 예상, 일반론은 금지하며, 실제 코멘트에 근거한 현상/영향/이슈 유형만 정리하세요. "
        "최종 출력은 JSON 객체 형식으로 반환해야 합니다."
    )

def build_user_prompt(df_final: pd.DataFrame, meta: Dict[str, Any]) -> str:
    """
    df_final의 주요 칼럼과 Fail 코멘트를 텍스트로 요약해 LLM에 전달.
    """
    lines = []
    lines.append("# META")
    for k, v in meta.items():
        if k == "metrics":
            continue
        lines.append(f"- {k}: {v}")
    lines.append("")

    # CLUSTERS_ISSUE_HW: 이슈 × GPU × Chipset × 해상도 군집 (count >= 2만 포함)
    metrics = meta.get("metrics", {}) if isinstance(meta, dict) else {}
    clusters = metrics.get("clusters", {}) if isinstance(metrics, dict) else {}
    issue_hw = clusters.get("issue_hw", []) or []

    strong_clusters = []
    for row in issue_hw:
        try:
            cnt = int(row.get("count", 0))
        except Exception:
            cnt = 0
        if cnt >= 2:
            strong_clusters.append(row)

    if strong_clusters:
        lines.append("# CLUSTERS_ISSUE_HW (count >= 2)")
        for row in strong_clusters:
            tag = row.get("feature_tag", "")
            hw = row.get("hw_value", "")
            res = row.get("resolution_group", "")
            cnt = row.get("count", "")
            repr_models = row.get("repr_models", [])
            if isinstance(repr_models, (list, tuple)):
                repr_models_str = ", ".join(map(str, repr_models))
            else:
                repr_models_str = str(repr_models)
            lines.append(
                f"- tag={tag}, hw={hw}, res={res}, count={cnt}, models=[{repr_models_str}]"
            )
        lines.append("")

    lines.append("# FAIL_ROWS")
    cols = []
    for c in ["Sheet","Checklist","Device(Model)","Rank","GPU","Chipset","RAM","OS","comment_text","issue_tags"]:
        if c in df_final.columns:
            cols.append(c)
    sample = df_final[cols].copy()

    for _, r in sample.iterrows():
        dev = _nkfc(r.get("Device(Model)", ""))
        gpu = _nkfc(r.get("GPU", ""))
        ch  = _nkfc(r.get("Chipset", ""))
        os_ = _nkfc(r.get("OS", ""))
        rank = _nkfc(r.get("Rank", ""))
        cmt = _nkfc(r.get("comment_text", ""))
        tags = r.get("issue_tags") or []
        tag_str = ", ".join(tags)
        lines.append(
            f"- Device={dev}, GPU={gpu}, Chipset={ch}, OS={os_}, Rank={rank}, Tags=[{tag_str}] / Comment={cmt}"
        )

    lines.append("")
    lines.append("# 작성 원칙 (개조식 문체 강제):\n"
        "1. 모든 문장은 군더더기 없는 **개조식(Bullet points)**으로 작성함.\n"
        "2. 종결 어미는 **명사형(~함, ~임, ~됨, ~확인됨)**으로 통일함.\n"
        "3. 배경 설명이나 접속사(따라서, 하지만 등)를 배제하고 핵심 팩트 위주로 기술함.\n\n")
    lines.append("# 요구 사항:")
    lines.append("1) 전체 테스트 결과를 기반으로 **종합 판정(Executive Summary)**를 작성된다.")
    lines.append(
        "2) Summary & Insight 항목은 문장 수나 길이에 제한은 없지만, 각 항목이 하나의 핵심 메시지/키워드만 담도록 "
        "짧고 명확하게 작성하며 불필요한 배경 설명은 쓰지 않는다."
    )
    lines.append(
        "3) Summary & Insight 항목들은 서로 다른 관점을 다루어야 한다 "
        "(예: 이슈 유형, 영향을 받는 GPU/Chipset/해상도, 사용자 영향, 품질/출시 리스크 등) "
        "및 이미 나온 내용과 거의 동일한 문장은 다시 작성하지 않는다(중복 금지)."
    )
    lines.append("4) Fail 유형 분석(실행 실패/크레시, UI 여백, UI 축소, 프레임 옵션 제한, iOS 노치/화면 이슈 등)을 실제 데이터가 있는 유형에 대해서만 정리한다.")
    lines.append("5) 최종 결론(출시 가능 여부, 선행 조치 필요 여부 등)을 데이터 기반으로만 작성한다.")
    lines.append("6) JSON 객체로 다음 키를 반드시 포함해 반환한다: ")
    lines.append('   { "executive_summary": string, "summary_insight": [string...], "issues": [...], "final_conclusion": string, "meta": {...} }')
    
    return "\n".join(lines)

# ==============================
# 4개 시트 리포트 작성
# ==============================
def write_excel_report(
    result: Dict[str, Any],
    df_final: pd.DataFrame,
    df_devices_all: pd.DataFrame,
    path: str
) -> None:
    try:
        import xlsxwriter
        engine = "xlsxwriter"
    except Exception:
        try:
            import openpyxl  # noqa
            engine = "openpyxl"
        except Exception:
            raise RuntimeError("엑셀 작성 엔진이 없습니다. `pip install xlsxwriter` 또는 `pip install openpyxl`")

    meta = result.get("meta", {}) if isinstance(result, dict) else {}
    metrics = meta.get("metrics", {}) if isinstance(meta, dict) else {}

    total_fail_issues = metrics.get("total_fail_issues", len(df_final))
    clusters = metrics.get("clusters", {}) if isinstance(metrics, dict) else {}
    by_issue_tag = metrics.get("by_issue_tag", [])
    clusters_feature = metrics.get("clusters_feature_detailed", [])
    clusters_issue_hw = clusters.get("issue_hw", [])

    executive_summary = result.get("executive_summary", "")
    raw_summary_insight = result.get("summary_insight", []) or []
    final_conclusion = result.get("final_conclusion", "")

    # Summary & Insight: 공백 정리 + 중복 제거 (개수/길이 제한 없음)
    summary_insight: List[str] = []
    seen = set()
    for s in raw_summary_insight:
        txt = re.sub(r"\s+", " ", str(s)).strip()
        if not txt:
            continue
        key = txt.lower()
        if key in seen:
            continue
        seen.add(key)
        summary_insight.append(txt)

    # 플랫폼 컬럼 추론
    df_devices = df_devices_all.copy()
    platform_col = None
    for cand in ["Platform", "platform", "OS_Type", "OS Type"]:
        if cand in df_devices.columns:
            platform_col = cand
            break
    if platform_col is None:
        df_devices["Platform"] = "AOS"
        platform_col = "Platform"

    # Device(Model) 보정
    def _norm_dev_key(x: Any) -> str:
        return str(x).strip().lower() if x is not None else ""

    if "Device(Model)" not in df_devices.columns:
        for cand in ["Device", "device", "Model", "MODEL", "단말명", "모델명"]:
            if cand in df_devices.columns:
                df_devices["Device(Model)"] = df_devices[cand]
                break

    if "Device(Model)" not in df_final.columns:
        for cand in ["Device", "device", "Model", "MODEL", "단말명", "모델명"]:
            if cand in df_final.columns:
                df_final["Device(Model)"] = df_final[cand]
                break

    # Fail/Crash 단말 키 수집
    fail_keys = set()
    if "Device(Model)" in df_final.columns:
        fail_keys = set(df_final["Device(Model)"].dropna().map(_norm_dev_key))

    crash_keys = set()
    if "Device(Model)" in df_final.columns and "issue_tags" in df_final.columns:
        for _, r in df_final.iterrows():
            tags = r.get("issue_tags") or []
            if "crash" in tags:
                crash_keys.add(_norm_dev_key(r.get("Device(Model)")))

    def _result_from_device_row(r):
        key = _norm_dev_key(r.get("Device(Model)"))
        if key in crash_keys:
            return "CRASH"
        if key in fail_keys:
            return "FAIL"
        return "PASS"

    df_devices["Result"] = df_devices.apply(_result_from_device_row, axis=1)

    # Fail Category / Detailed Issue 매핑
    tag_map: Dict[str, str] = {}
    comment_map: Dict[str, str] = {}
    if "Device(Model)" in df_final.columns:
        for _, r in df_final.iterrows():
            k = _norm_dev_key(r.get("Device(Model)"))
            tags = r.get("issue_tags") or []
            if tags and k not in tag_map:
                tag_map[k] = tags[0]
            cmt = str(r.get("comment_text") or "").strip()
            if cmt and k not in comment_map:
                comment_map[k] = cmt

    def _fail_category_from_device_row(r):
        key = _norm_dev_key(r.get("Device(Model)"))
        tag = tag_map.get(key, "")
        if not tag:
            return ""
        return TAG_LABEL_KO.get(tag, tag)

    def _detailed_issue_from_device_row(r):
        key = _norm_dev_key(r.get("Device(Model)"))
        return comment_map.get(key, "")

    df_devices["Fail Category"] = df_devices.apply(_fail_category_from_device_row, axis=1)
    df_devices["Detailed Issue"] = df_devices.apply(_detailed_issue_from_device_row, axis=1)
    df_devices["Action Required"] = ""

    # 플랫폼별 통계
    platform_stats = []
    for plat in sorted(df_devices[platform_col].dropna().unique().tolist()):
        sub = df_devices[df_devices[platform_col] == plat]
        total = len(sub)
        pass_cnt = int((sub["Result"] == "PASS").sum())
        fail_cnt = int((sub["Result"] == "FAIL").sum())
        crash_cnt = int((sub["Result"] == "CRASH").sum())
        platform_stats.append({
            "Platform": plat,
            "Total": total,
            "Pass": pass_cnt,
            "Fail": fail_cnt,
            "Crash": crash_cnt,
        })
    df_platform_stats = pd.DataFrame(platform_stats)

    with pd.ExcelWriter(path, engine=engine) as wr:
        # 공통: 열 너비 자동 맞춤 + 자동 줄바꿈
        def _autofit_and_wrap(sheet_name: str, df: pd.DataFrame, wrap_all: bool = True):
            """
            df를 쓴 뒤, 해당 시트의 열 너비를 데이터 길이에 맞게 조정하고
            텍스트 자동 줄바꿈을 적용한다.
            """
            if sheet_name not in wr.sheets:
                return
            ws = wr.sheets[sheet_name]

            # xlsxwriter 엔진인 경우
            book = getattr(wr, "book", None)
            if book is not None:
                wrap_fmt = book.add_format({"text_wrap": True, "valign": "top"})
                for col_idx, col_name in enumerate(df.columns):
                    series = df[col_name].astype(str)
                    max_len = max(series.map(len).max(), len(str(col_name)))
                    width = min(max_len + 2, 60)
                    if wrap_all:
                        ws.set_column(col_idx, col_idx, width, wrap_fmt)
                    else:
                        ws.set_column(col_idx, col_idx, width)
                return

            # openpyxl 엔진인 경우: column_dimensions 이용
            try:
                from openpyxl.utils import get_column_letter
            except Exception:
                return

            for col_idx, col_name in enumerate(df.columns):
                col_letter = get_column_letter(col_idx + 1)
                series = df[col_name].astype(str)
                max_len = max(series.map(len).max(), len(str(col_name)))
                width = min(max_len + 2, 60)
                ws.column_dimensions[col_letter].width = width

        # Sheet1: Summary
        summary_rows = []
        summary_rows.append({"Section": "테스트 개요", "Item": "빌드 버전", "Value": meta.get("build_version", "KP 4.2.0")})
        summary_rows.append({"Section": "테스트 개요", "Item": "테스트 범위", "Value": meta.get("scope", "Android / iOS 호환성 검증")})
        summary_rows.append({"Section": "테스트 개요", "Item": "테스트 단말 수", "Value": f"{len(df_devices)}대 (AOS/iOS 합산)"})
        summary_rows.append({"Section": "테스트 개요", "Item": "총 Fail 이슈 수", "Value": total_fail_issues})
        df_summary_head = pd.DataFrame(summary_rows)

        summary_block = []
        if executive_summary:
            lines = [ln.strip() for ln in str(executive_summary).split("\n") if ln.strip()]
            for i, ln in enumerate(lines, start=1):
                summary_block.append({"Section": "종합 판정", "No": i, "Text": ln})

        for i, ln in enumerate(summary_insight, start=1):
            summary_block.append({"Section": "Summary & Insight", "No": i, "Text": str(ln).strip()})

        if final_conclusion:
            lines = [ln.strip() for ln in str(final_conclusion).split("\n") if ln.strip()]
            for i, ln in enumerate(lines, start=1):
                summary_block.append({"Section": "최종 결론", "No": i, "Text": ln})

        df_summary_block = pd.DataFrame(summary_block)

        start_row = 0
        df_summary_head.to_excel(wr, sheet_name="Summary", index=False, startrow=start_row)
        start_row += len(df_summary_head) + 2

        if not df_platform_stats.empty:
            df_platform_stats.to_excel(wr, sheet_name="Summary", index=False, startrow=start_row)
            start_row += len(df_platform_stats) + 2

        if not df_summary_block.empty:
            df_summary_block.to_excel(wr, sheet_name="Summary", index=False, startrow=start_row)

        # Summary 시트 자동 맞춤 (head + platform_stats + block 합쳐서 폭 계산)
        df_summary_for_width = pd.concat(
            [df_summary_head, df_platform_stats, df_summary_block],
            ignore_index=True
        ).fillna("")
        _autofit_and_wrap("Summary", df_summary_for_width, wrap_all=True)

        # Sheet2: Device_List
        base_cols = ["Sheet", "Platform", "Device(Model)", "GPU", "Chipset", "RAM", "OS", "Rank"]
        dev_cols = [c for c in base_cols if c in df_devices.columns]
        cols_out = dev_cols + ["Result", "Fail Category", "Detailed Issue", "Action Required"]
        df_devices[cols_out].to_excel(wr, sheet_name="Device_List", index=False)
        _autofit_and_wrap("Device_List", df_devices[cols_out], wrap_all=True)

        # Sheet3: Final_Report
        fr_rows = []
        fr_rows.append({"Section": "1. 테스트 개요", "Content": f"빌드 버전: {meta.get('build_version', 'KP 4.2.0')}"})
        fr_rows.append({"Section": "1. 테스트 개요", "Content": f"테스트 범위: {meta.get('scope', 'Android / iOS 호환성 검증')}"})
        fr_rows.append({"Section": "1. 테스트 개요", "Content": f"테스트 단말 수: {len(df_devices)}대 (AOS/iOS 합산)"})

        fr_rows.append({"Section": "2. 테스트 단말 구성", "Content": "플랫폼별 Pass/Fail/Crash 요약은 아래 표 참조."})

        if executive_summary:
            lines = [ln.strip() for ln in str(executive_summary).split("\n") if ln.strip()]
            for ln in lines:
                fr_rows.append({"Section": "3. 종합 판정", "Content": ln})

        for ln in summary_insight:
            fr_rows.append({"Section": "4. Summary & Insight", "Content": str(ln).strip()})

        issues_list = result.get("issues", []) or []
        for iss in issues_list:
            title = iss.get("title", "")
            desc = iss.get("description", "")
            fr_rows.append({"Section": "5. Fail 유형 분석", "Content": f"[{title}] {desc}"})

        if final_conclusion:
            lines = [ln.strip() for ln in str(final_conclusion).split("\n") if ln.strip()]
            for ln in lines:
                fr_rows.append({"Section": "7. 최종 결론", "Content": ln})

        df_fr = pd.DataFrame(fr_rows)
        df_fr.to_excel(wr, sheet_name="Final_Report", index=False)

        if not df_platform_stats.empty:
            start_row_fr = len(df_fr) + 3
            df_platform_stats.to_excel(wr, sheet_name="Final_Report", index=False, startrow=start_row_fr)

        # Final_Report 시트 자동 맞춤
        _autofit_and_wrap("Final_Report", df_fr, wrap_all=True)

        # Sheet4: Clusters (by_issue_tag / feature_detailed / issue_hw 등 통합)
        cluster_rows = []

        by_gpu = clusters.get("by_gpu", []) or []
        for row in by_gpu:
            cluster_rows.append({
                "cluster_type": "by_gpu",
                "feature_tag": "",
                "hw_type": "gpu",
                "hw_value": row.get("GPU", row.get("gpu", "")),
                "resolution_group": "",
                "count": row.get("count", row.get("Count", "")),
                "extra": "",
            })

        by_chipset = clusters.get("by_chipset", []) or []
        for row in by_chipset:
            cluster_rows.append({
                "cluster_type": "by_chipset",
                "feature_tag": "",
                "hw_type": "chipset",
                "hw_value": row.get("Chipset", row.get("chipset", "")),
                "resolution_group": "",
                "count": row.get("count", row.get("Count", "")),
                "extra": "",
            })

        for row in by_issue_tag or []:
            val = row.get("value", "")
            cluster_rows.append({
                "cluster_type": "by_issue_tag",
                "feature_tag": val,
                "hw_type": "",
                "hw_value": "",
                "resolution_group": "",
                "count": row.get("count", ""),
                "extra": "",
            })

        for row in clusters_feature or []:
            tag = row.get("feature_tag", "")
            c = row.get("count", "")
            repr_models = ", ".join(row.get("repr_models", []) or [])
            cluster_rows.append({
                "cluster_type": "feature_detailed",
                "feature_tag": tag,
                "hw_type": "",
                "hw_value": "",
                "resolution_group": "",
                "count": c,
                "extra": repr_models,
            })

        for row in clusters_issue_hw or []:
            tag = row.get("feature_tag", "")
            hw_type = row.get("hw_type", "")
            hw_value = row.get("hw_value", "")
            resg = row.get("resolution_group", "")
            c = row.get("count", "")
            repr_models = ", ".join(row.get("repr_models", []) or [])
            cluster_rows.append({
                "cluster_type": "issue_hw",
                "feature_tag": tag,
                "hw_type": hw_type,
                "hw_value": hw_value,
                "resolution_group": resg,
                "count": c,
                "extra": repr_models,
            })

        df_clusters_all = pd.DataFrame(cluster_rows)
        if df_clusters_all.empty:
            df_clusters_all = pd.DataFrame(
                [{"cluster_type": "(none)", "feature_tag": "", "hw_type": "", "hw_value": "",
                  "resolution_group": "", "count": "", "extra": ""}]
            )

        df_clusters_all.to_excel(wr, sheet_name="Clusters", index=False)
        _autofit_and_wrap("Clusters", df_clusters_all, wrap_all=True)
